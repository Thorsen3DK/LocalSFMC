"""
Excel Data Loader — Reads .xlsx files as Data Extensions.

Each sheet in an Excel file becomes a named Data Extension.
First row = column headers (field names), remaining rows = records.

DE Name Mapping:
  A de_mapping.json file in the data_extensions folder maps AMPScript
  DE names to actual Excel sheet names. Example:
    {"DINGO_SUBS": "4-19-2026-DE_SUBS_FOR_GEN_VARSL"}
"""

from __future__ import annotations

import csv
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

MAPPING_FILE = "de_mapping.json"

# Strips leading date prefixes like "4-28-2026-" or "04-28-2026-" from a name,
# so an uploaded file named "4-28-2026-WOMBAT_SUBS.csv" is also reachable as
# "WOMBAT_SUBS" (the bare name AMPScript templates typically reference).
_DATE_PREFIX_RE = re.compile(r"^\d{1,2}-\d{1,2}-\d{2,4}-")

# Cache for load_all() keyed on (dir, frozenset of (path, mtime, size))
# When any CSV/XLSX/mapping file changes, the signature changes and the cache misses.
# Critically, cache hits return the SAME list objects, so LookupRows hash indexes
# (keyed by id(rows) in ampscript.functions) remain valid across requests.
_LOAD_ALL_CACHE: Dict[tuple, Dict[str, List[Dict[str, Any]]]] = {}


def _dir_signature(de_dir: Path) -> tuple:
    parts = []
    for f in sorted(de_dir.iterdir()):
        if f.name.startswith("~$"):
            continue
        if f.suffix.lower() in (".csv", ".xlsx", ".json"):
            try:
                st = f.stat()
                parts.append((f.name, st.st_mtime_ns, st.st_size))
            except FileNotFoundError:
                pass
    return tuple(parts)


def load_excel_file(filepath: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Load a single .xlsx file. Returns a dict mapping sheet names to lists of row dicts.

    Example:
        {"Sheet1": [{"FirstName": "Alice", "Email": "alice@ex.com"}, ...]}
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    result: Dict[str, List[Dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue

        headers = [str(h).strip() if h is not None else f"Column{i}"
                   for i, h in enumerate(rows[0], 1)]
        records = []
        for row in rows[1:]:
            # Skip entirely empty rows
            if all(cell is None for cell in row):
                continue
            record = {}
            for header, cell in zip(headers, row):
                record[header] = _normalize_cell(cell)
            records.append(record)
        result[sheet_name] = records

    wb.close()
    return result


def load_csv_file(filepath: str) -> List[Dict[str, Any]]:
    """
    Load a single .csv file. Returns a list of row dicts.
    """
    records = []
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            record = {k.strip(): _normalize_cell(v.strip() if isinstance(v, str) else v)
                      for k, v in row.items() if k is not None}
            records.append(record)
    return records


def load_all(data_extensions_dir: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Load all .xlsx and .csv files from a directory.

    Returns a dict mapping DE names to row lists. DE name = sheet name if unique,
    otherwise "filename_sheetname". Also applies aliases from de_mapping.json.

    Result is cached across calls and invalidated when any source file's mtime/size
    changes. Returns the same underlying list objects on cache hits, so downstream
    indexes (e.g., LookupRows) remain valid.
    """
    de_dir = Path(data_extensions_dir)
    if not de_dir.is_dir():
        return {}

    cache_key = (str(de_dir.resolve()), _dir_signature(de_dir))
    cached = _LOAD_ALL_CACHE.get(cache_key)
    if cached is not None:
        return cached

    all_des: Dict[str, List[Dict[str, Any]]] = {}
    seen_sheet_names: Dict[str, int] = {}

    # First pass: count sheet name occurrences across xlsx files
    file_sheets: List[tuple] = []
    for xlsx_file in sorted(de_dir.glob("*.xlsx")):
        if xlsx_file.name.startswith("~$"):
            continue  # skip temp files
        sheets = load_excel_file(str(xlsx_file))
        for sheet_name, records in sheets.items():
            file_sheets.append((xlsx_file.stem, sheet_name, records))
            seen_sheet_names[sheet_name] = seen_sheet_names.get(sheet_name, 0) + 1

    # Second pass: assign names for xlsx
    for file_stem, sheet_name, records in file_sheets:
        if seen_sheet_names[sheet_name] > 1:
            de_name = f"{file_stem}_{sheet_name}"
        else:
            de_name = sheet_name
        all_des[de_name] = records

    # Load CSV files (DE name = file stem)
    for csv_file in sorted(de_dir.glob("*.csv")):
        if csv_file.name.startswith("~$"):
            continue
        try:
            records = load_csv_file(str(csv_file))
            all_des[csv_file.stem] = records
        except Exception:
            pass

    # Auto-register date-prefix-stripped aliases. e.g. "4-28-2026-WOMBAT_SUBS"
    # is also exposed as "WOMBAT_SUBS" so templates can use the bare DE name.
    # Skip when the stripped name collides with another DE or another stripped name.
    stripped_names: Dict[str, str] = {}
    for de_name in list(all_des.keys()):
        stripped = _DATE_PREFIX_RE.sub("", de_name)
        if stripped == de_name or not stripped:
            continue
        if stripped in all_des:
            continue  # don't shadow an explicit DE
        if stripped in stripped_names:
            stripped_names[stripped] = ""  # collision: drop both
            continue
        stripped_names[stripped] = de_name
    for alias, source in stripped_names.items():
        if source and alias not in all_des:
            all_des[alias] = all_des[source]

    # Apply DE name aliases from mapping file (overrides auto-aliases)
    mapping = load_de_mapping(data_extensions_dir)
    for alias, target in mapping.items():
        target_lower = target.lower()
        for de_name, records in list(all_des.items()):
            if de_name.lower() == target_lower:
                all_des[alias] = records
                break

    _LOAD_ALL_CACHE.clear()  # only keep latest signature; data dirs evolve over time
    _LOAD_ALL_CACHE[cache_key] = all_des
    return all_des


def get_send_list(
    data_extensions_dir: str, filename: str, sheet: Optional[str] = None
) -> List[Dict[str, Any]]:
    """Load a specific Excel or CSV file as the send list (subscriber rows)."""
    filepath = os.path.join(data_extensions_dir, filename)

    if filename.lower().endswith('.csv'):
        return load_csv_file(filepath)

    sheets = load_excel_file(filepath)

    if sheet and sheet in sheets:
        return sheets[sheet]

    # Default to first sheet
    if sheets:
        return next(iter(sheets.values()))

    return []


def list_excel_files(data_extensions_dir: str) -> List[Dict[str, Any]]:
    """
    List all available Excel files and their sheets/row counts.

    Returns: [{"filename": "data.xlsx", "sheets": [{"name": "Sheet1", "rows": 100}, ...]}]
    """
    de_dir = Path(data_extensions_dir)
    if not de_dir.is_dir():
        return []

    result = []
    for xlsx_file in sorted(de_dir.glob("*.xlsx")):
        if xlsx_file.name.startswith("~$"):
            continue
        try:
            sheets = load_excel_file(str(xlsx_file))
            sheet_info = [{"name": name, "rows": len(rows)} for name, rows in sheets.items()]
            result.append({"filename": xlsx_file.name, "sheets": sheet_info})
        except Exception:
            result.append({"filename": xlsx_file.name, "sheets": [], "error": True})
    return result


def _normalize_cell(val: Any) -> Any:
    """Normalize cell values — keep native types but convert None to empty string."""
    if val is None:
        return ""
    return val


# ---------------------------------------------------------------------------
# DE Name Mapping
# ---------------------------------------------------------------------------

def load_de_mapping(data_extensions_dir: str) -> Dict[str, str]:
    """Load DE name mapping from de_mapping.json. Returns {alias: actual_sheet_name}."""
    mapping_path = os.path.join(data_extensions_dir, MAPPING_FILE)
    if os.path.isfile(mapping_path):
        with open(mapping_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_de_mapping(data_extensions_dir: str, mapping: Dict[str, str]):
    """Save DE name mapping to de_mapping.json."""
    mapping_path = os.path.join(data_extensions_dir, MAPPING_FILE)
    with open(mapping_path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)


def list_all_de_names(data_extensions_dir: str) -> List[str]:
    """List all available DE names (sheet names from all Excel files)."""
    de_dir = Path(data_extensions_dir)
    if not de_dir.is_dir():
        return []
    names = set()
    for xlsx_file in sorted(de_dir.glob("*.xlsx")):
        if xlsx_file.name.startswith("~$"):
            continue
        try:
            sheets = load_excel_file(str(xlsx_file))
            names.update(sheets.keys())
        except Exception:
            pass
    return sorted(names)
